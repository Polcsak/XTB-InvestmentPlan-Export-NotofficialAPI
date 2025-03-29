import os
import time
import shutil
import csv
import re
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
import subprocess
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# ----------------------------
# Configuration Settings
# ----------------------------
chrome_path = r"path/to/chrome.exe  # path to chrome.exe"
chrome_debug_port = "9222"
user_data_path = r"path/to/chrome_profile  # path to the Chrome profile folder"
csv_file_path = r"path/to/your/csv_file.csv  # path to the CSV file"

# ----------------------------
# Step 1: Launch Chrome in Debug Mode with a Specific Test Profile
# ----------------------------
chrome_command = f'"{chrome_path}" --remote-debugging-port={chrome_debug_port} --user-data-dir="{user_data_path}" --disable-component-update'
subprocess.Popen(chrome_command, shell=True)
time.sleep(3)

# ----------------------------
# Step 2: Connect to the Running Chrome Instance and Load the XTB Website
# ----------------------------
options = webdriver.ChromeOptions()
options.debugger_address = f"127.0.0.1:{chrome_debug_port}"
driver = webdriver.Chrome(options=options)
driver.get("https://xstation5.xtb.com")
time.sleep(5)

# ----------------------------
# Step 3: Click on the "Login" Button
# ----------------------------
login_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, "input.xs-btn-ok-login"))
)
login_button.click()
print("Logged in successfully (if the credentials were correct).")
time.sleep(10)

# ----------------------------
# Step 4: Click on the First Export Button
# ----------------------------
try:
    export_button = driver.execute_script('''
        return document
            .querySelector("#contentLayer0 > div.xs-hdivided-container > div:nth-child(2) > div.xs-tab-module-container.xs-tab-module-container-active > div > div > xs6-history")
            .shadowRoot.querySelector("xs6-history-feature > div > div.tab-bar-container > button");
    ''')
    
    if export_button:
        driver.execute_script("arguments[0].click();", export_button)
        print("Clicked on the Export button.")
    else:
        print("Export button not found.")
except Exception as e:
    print("Error clicking the Export button:", e)

# ----------------------------
# Step 5: Click on the Date Input Field
# ----------------------------
time.sleep(2)
try:
    date_input = driver.execute_script('''
        return document.querySelector("#contentLayer0 > div.xs-hdivided-container > div:nth-child(2) > div.xs-tab-module-container.xs-tab-module-container-active > div > div > xs6-history")
            .shadowRoot.querySelector("xs6-history-feature > div > div.history-tabs-container > div > xs6-history_featureclosedpositions")
            .shadowRoot.querySelector("xs6-closed-positions-feature > div > xs6-export-report-dialog > dialog > div > section > div.pds-modal__content > div:nth-child(1) > xs6-date-select > div > pds-datepicker-input-from > pds-control > div > input");
    ''')
    
    driver.execute_script("arguments[0].click();", date_input)
    print("Clicked on the date input field.")
except Exception as e:
    print("Error clicking the date input field:", e)

# ----------------------------
# Step 6: Click on the "All" Records Option
# ----------------------------
time.sleep(2)
try:
    all_option = driver.execute_script('''
        return document.querySelector("#contentLayer0 > div.xs-hdivided-container > div:nth-child(2) > div.xs-tab-module-container.xs-tab-module-container-active > div > div > xs6-history")
            .shadowRoot.querySelector("xs6-history-feature > div > div.history-tabs-container > div > xs6-history_featureclosedpositions")
            .shadowRoot.querySelector("#cdk-menu-0 > pds-action-list-item:nth-child(7)");
    ''')
    
    if all_option:
        driver.execute_script("arguments[0].click();", all_option)
        print("Clicked on the 'All' records option.")
    else:
        print("'All' records option not found.")
except Exception as e:
    print("Error clicking the 'All' records option:", e)

# ----------------------------
# Step 7: Click on the Final Export Button (Export Report)
# ----------------------------
time.sleep(2)
try:
    export_button_final = driver.execute_script('''
        return document.querySelector("#contentLayer0 > div.xs-hdivided-container > div:nth-child(2) > div.xs-tab-module-container.xs-tab-module-container-active > div > div > xs6-history")
            .shadowRoot.querySelector("xs6-history-feature > div > div.history-tabs-container > div > xs6-history_featureclosedpositions")
            .shadowRoot.querySelector("xs6-closed-positions-feature > div > xs6-export-report-dialog > dialog > div > footer > div > button.pds-button.pds-button--size--l.pds-button--style--primary");
    ''')
    if export_button_final:
        driver.execute_script("arguments[0].click();", export_button_final)
        print("Clicked on the Export Report button.")
    else:
        print("Final Export button not found.")
except Exception as e:
    print("Error clicking the final Export button:", e)

# ----------------------------
# Step 8: Wait for the Download and Close the Automated Chrome Instance
# ----------------------------
time.sleep(15)
# Terminate processes launched with the specified Chrome profile
os.system('wmic process where "commandline like \'%%chrome_selenium_profile%%\'" call terminate >nul 2>&1')
print("Automated Chrome instance has been closed.")

# ----------------------------
# Step 9: Move the Downloaded File and Rename the Worksheet in the Excel File
# ----------------------------
downloads_folder = r"path/to/downloads_folder  # path to the downloads folder"
target_folder = r"path/to/target_folder  # path to the target folder"

time.sleep(3)

# Locate and move the file
for filename in os.listdir(downloads_folder):
    if filename.startswith("account_50266286") and filename.endswith(".xlsx"):
        source_path = os.path.join(downloads_folder, filename)
        target_path = os.path.join(target_folder, "xtb_export.xlsx")
        shutil.move(source_path, target_path)
        print(f"File moved and renamed to: {target_path}")

        # Rename the worksheet starting with "OPEN POSITION" to "ExportXTB"
        wb = load_workbook(target_path)
        sheet_renamed = False
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("OPEN POSITION"):
                wb[sheet_name].title = "ExportXTB"
                wb.save(target_path)
                print(f"Worksheet '{sheet_name}' renamed to: ExportXTB")
                sheet_renamed = True
                break
        if not sheet_renamed:
            print("Worksheet starting with 'OPEN POSITION' was not found.")
        break
else:
    print("File not found in the Downloads folder.")
